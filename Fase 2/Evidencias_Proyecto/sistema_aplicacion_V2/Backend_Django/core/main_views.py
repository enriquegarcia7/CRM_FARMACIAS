from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.db.models import Count, Sum, Q, F, Max
from django.db.models.functions import TruncMonth, TruncDate, Coalesce
from django.utils import timezone
from datetime import timedelta
from .models import (
    Cliente, Transaccion, Producto, Proveedor,
    OfertaLaboratorio, SugerenciaCompra, Venta, DetalleVenta
)
from .serializers import (
    ClienteSerializer, TransaccionSerializer, ProductoSerializer,
    ProductoListSerializer,
    ProveedorSerializer, OfertaLaboratorioSerializer,
    OfertaLaboratorioDetalladaSerializer,
    SugerenciaCompraSerializer, VentaSerializer
)


class OfertasPagination(PageNumberPagination):
    """Paginación optimizada para ofertas - 50 items por página"""
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 100


class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    pagination_class = OfertasPagination  # 50 clientes por página
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'rut', 'correo', 'telefono']
    ordering_fields = ['nombre', 'rut', 'correo', 'total_ventas', 'monto_total', 'ultima_compra']
    ordering = ['nombre']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Anotar campos calculados para permitir ordenamiento
        # Usamos 'total_ventas' en lugar de 'total_compras' para evitar colisión con la property del modelo
        from django.db.models import Value
        from decimal import Decimal

        queryset = queryset.annotate(
            total_ventas=Count('ventas', filter=Q(ventas__estado__codigo='completada')),
            monto_total=Coalesce(Sum('ventas__total', filter=Q(ventas__estado__codigo='completada')), Value(Decimal('0'))),
            ultima_compra=Max('ventas__fecha', filter=Q(ventas__estado__codigo='completada'))
        )

        # Búsqueda personalizada con normalización de RUT
        search = self.request.query_params.get('search', '')
        if search:
            search_normalized = search.replace('.', '').replace('-', '').strip()
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(rut__icontains=search_normalized) |
                Q(correo__icontains=search)
            )

        # Filtro por tipo de cliente
        tipo = self.request.query_params.get('tipo', '')
        if tipo == 'frecuentes':
            queryset = queryset.filter(total_ventas__gte=5)
        elif tipo == 'normales':
            queryset = queryset.filter(total_ventas__lt=5)

        return queryset

    def list(self, request, *args, **kwargs):
        if request.query_params.get('search'):
            self.filter_backends = [filters.OrderingFilter]
        return super().list(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def frecuentes(self, request):
        """Obtiene clientes frecuentes (con 5 o más compras) con paginación"""
        clientes = Cliente.objects.annotate(
            total_ventas=Count('ventas', filter=Q(ventas__estado__codigo='completada'))
        ).filter(total_ventas__gte=5).order_by('-total_ventas', '-id')

        # Aplicar paginación
        page = self.paginate_queryset(clientes)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(clientes, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Obtiene estadísticas globales de clientes con validación de correo en BD"""
        total_clientes = Cliente.objects.count()

        # Anotar todos los clientes con su total de compras
        clientes_anotados = Cliente.objects.annotate(
            total_compras=Count('ventas', filter=Q(ventas__estado__codigo='completada'))
        )

        # Clientes frecuentes (>= 5 compras)
        frecuentes_queryset = clientes_anotados.filter(total_compras__gte=5)
        clientes_frecuentes = frecuentes_queryset.count()

        # Clientes normales (< 5 compras)
        clientes_normales = clientes_anotados.filter(total_compras__lt=5).count()

        # Elegibles: frecuentes con correo válido (filtrado directo en PostgreSQL)
        # Regex PostgreSQL: tiene @, tiene . después del @, mínimo 2 chars después del último .
        elegibles_ofertas = frecuentes_queryset.filter(
            correo__isnull=False,
            correo__regex=r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        ).exclude(correo='').count()

        # Calcular sin correo y con correo inválido
        con_correo = frecuentes_queryset.filter(correo__isnull=False).exclude(correo='').count()
        sin_correo = clientes_frecuentes - con_correo
        correo_invalido = con_correo - elegibles_ofertas

        return Response({
            'total_clientes': total_clientes,
            'clientes_frecuentes': clientes_frecuentes,
            'clientes_normales': clientes_normales,
            'elegibles_ofertas': elegibles_ofertas,
            'sin_correo': sin_correo,
            'correo_invalido': correo_invalido
        })

    @action(detail=True, methods=['get'])
    def productos_comprados(self, request, pk=None):
        """Obtiene los productos más comprados por un cliente específico"""
        cliente = self.get_object()

        # Obtener productos comprados por el cliente desde DetalleVenta
        productos = DetalleVenta.objects.filter(
            venta__cliente=cliente,
            venta__estado__codigo='completada'
        ).values(
            'producto__id',
            'producto__codigo',
            'producto__nombre',
            'producto__descripcion'
        ).annotate(
            total_cantidad=Sum('cantidad'),
            total_monto=Sum('subtotal')
        ).order_by('-total_cantidad')[:10]  # Top 10 productos más comprados

        return Response({
            'cliente_id': cliente.id,
            'cliente_nombre': cliente.nombre,
            'productos': list(productos)
        })

    @action(detail=True, methods=['post'], url_path='enviar-oferta')
    def enviar_oferta(self, request, pk=None):
        """
        Envía una oferta por correo a un cliente específico.

        Body:
            - descuento: Porcentaje de descuento (3, 5, 10, 15, 20)
            - productos: Lista de nombres de productos (opcional, si no se envía se usan los más comprados)
        """
        cliente = self.get_object()

        # Validar que el cliente tenga correo
        if not cliente.correo:
            return Response({
                'success': False,
                'error': 'El cliente no tiene correo electrónico registrado'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Obtener datos del request
        descuento = request.data.get('descuento', 10)

        # Obtener productos más comprados con sus precios
        productos_comprados = DetalleVenta.objects.filter(
            venta__cliente=cliente,
            venta__estado__codigo='completada'
        ).values(
            'producto__nombre',
            'producto__precio_venta'
        ).annotate(
            total_cantidad=Sum('cantidad')
        ).order_by('-total_cantidad')[:5]

        # Construir lista de productos con precios
        productos_con_precio = []
        for p in productos_comprados:
            precio = float(p['producto__precio_venta']) if p['producto__precio_venta'] else 0
            productos_con_precio.append({
                'nombre': p['producto__nombre'],
                'precio': precio
            })

        # Si no hay productos, usar genérico
        if not productos_con_precio:
            productos_con_precio = [{'nombre': 'Productos de su preferencia', 'precio': 0}]

        try:
            from core.services.gmail_service import GmailService

            gmail = GmailService()
            result = gmail.send_offer_email(
                cliente_nombre=cliente.nombre,
                cliente_email=cliente.correo,
                productos=productos_con_precio,
                descuento_porcentaje=descuento
            )

            if result:
                return Response({
                    'success': True,
                    'message': f'Oferta enviada exitosamente a {cliente.correo}',
                    'message_id': result.get('id'),
                    'cliente': cliente.nombre,
                    'descuento': descuento,
                    'productos_incluidos': [p['nombre'] for p in productos_con_precio]
                })
            else:
                return Response({
                    'success': False,
                    'error': 'Error al enviar el correo'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except FileNotFoundError as e:
            return Response({
                'success': False,
                'error': 'Gmail no está autenticado. Por favor autentique Gmail desde la sección ETL.'
            }, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post', 'options'], url_path='enviar-ofertas-masivas')
    def enviar_ofertas_masivas(self, request):
        """
        Envía ofertas a clientes frecuentes elegibles usando streaming SSE.
        Filtrado rápido directo en PostgreSQL.

        Excluye automáticamente:
        - Clientes sin correo válido
        - Clientes sin nombre o con nombre genérico (Cliente_XXXX)
        - Clientes sin RUT

        Body:
            - descuento: Porcentaje de descuento (3, 5, 10, 15, 20)
        """
        from django.http import StreamingHttpResponse, HttpResponse
        import json

        # Manejar preflight CORS (OPTIONS request)
        if request.method == 'OPTIONS':
            response = HttpResponse()
            response['Access-Control-Allow-Origin'] = '*'
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type'
            return response

        descuento = request.data.get('descuento', 10)
        hace_24h = timezone.now() - timedelta(hours=24)

        # Query base: frecuentes con correo válido, nombre real y RUT (filtrado directo en PostgreSQL)
        frecuentes_base = Cliente.objects.annotate(
            total_ventas=Count('ventas', filter=Q(ventas__estado__codigo='completada'))
        ).filter(
            total_ventas__gte=5,
            # Correo válido
            correo__isnull=False,
            correo__regex=r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$',
            # Nombre válido (no vacío)
            nombre__isnull=False,
            # RUT válido (no vacío)
            rut__isnull=False
        ).exclude(
            correo=''
        ).exclude(
            nombre=''
        ).exclude(
            rut=''
        ).exclude(
            # Excluir nombres genéricos: "Cliente XXXXX" o "Cliente_XXXXX"
            nombre__regex=r'^Cliente[\s_]?\d+$'
        ).exclude(
            nombre__iexact='Cliente General'
        )

        total_frecuentes = Cliente.objects.annotate(
            total_ventas=Count('ventas', filter=Q(ventas__estado__codigo='completada'))
        ).filter(total_ventas__gte=5).count()

        # Estadísticas rápidas
        total_elegibles = frecuentes_base.count()
        ya_enviados_24h = frecuentes_base.filter(ultima_oferta_enviada__gte=hace_24h).count()

        # Clientes a enviar: elegibles que no han recibido en 24h
        clientes_frecuentes = list(frecuentes_base.filter(
            Q(ultima_oferta_enviada__isnull=True) | Q(ultima_oferta_enviada__lt=hace_24h)
        ))

        # Calcular excluidos detalladamente
        frecuentes_todos = Cliente.objects.annotate(
            total_ventas=Count('ventas', filter=Q(ventas__estado__codigo='completada'))
        ).filter(total_ventas__gte=5)

        # Sin correo
        sin_correo = frecuentes_todos.filter(
            Q(correo__isnull=True) | Q(correo='')
        ).count()

        # Con correo inválido (tiene correo pero no cumple regex)
        con_correo = frecuentes_todos.exclude(correo__isnull=True).exclude(correo='')
        correos_invalidos = con_correo.exclude(
            correo__regex=r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        ).count()

        # Sin nombre real (vacío o genérico)
        sin_nombre_real = frecuentes_todos.filter(
            Q(nombre__isnull=True) |
            Q(nombre='') |
            Q(nombre__regex=r'^Cliente[\s_]?\d+$') |
            Q(nombre__iexact='Cliente General')
        ).count()

        # Total excluidos por datos incompletos
        datos_incompletos = total_frecuentes - total_elegibles - ya_enviados_24h

        if not clientes_frecuentes:
            return Response({
                'success': False,
                'error': 'No hay clientes elegibles pendientes.',
                'detalle': {
                    'total_frecuentes': total_frecuentes,
                    'elegibles': total_elegibles,
                    'sin_correo': sin_correo,
                    'correo_invalido': correos_invalidos,
                    'sin_nombre_real': sin_nombre_real,
                    'ya_enviados_24h': ya_enviados_24h
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        def generate_progress():
            """Generador que envía progreso en formato SSE con envío paralelo"""
            from core.services.gmail_service import GmailService
            from concurrent.futures import ThreadPoolExecutor, as_completed
            import threading

            try:
                gmail = GmailService()
            except FileNotFoundError:
                yield f"data: {json.dumps({'type': 'error', 'message': 'Gmail no está autenticado'})}\n\n"
                return

            total = len(clientes_frecuentes)
            enviados = 0
            errores = []
            lock = threading.Lock()

            # Enviar evento inicial con información de excluidos
            yield f"data: {json.dumps({'type': 'start', 'total': total, 'total_frecuentes': total_frecuentes, 'sin_correo': sin_correo, 'correo_invalido': correos_invalidos, 'sin_nombre_real': sin_nombre_real, 'ya_enviados_24h': ya_enviados_24h})}\n\n"

            # Pre-cargar productos de todos los clientes para evitar queries en threads
            clientes_con_productos = []
            for cliente in clientes_frecuentes:
                productos_comprados = DetalleVenta.objects.filter(
                    venta__cliente=cliente,
                    venta__estado__codigo='completada'
                ).values(
                    'producto__nombre',
                    'producto__precio_venta'
                ).annotate(
                    total_cantidad=Sum('cantidad')
                ).order_by('-total_cantidad')[:5]

                productos_con_precio = []
                for p in productos_comprados:
                    precio = float(p['producto__precio_venta']) if p['producto__precio_venta'] else 0
                    productos_con_precio.append({
                        'nombre': p['producto__nombre'],
                        'precio': precio
                    })

                if not productos_con_precio:
                    productos_con_precio = [{'nombre': 'Productos de su preferencia', 'precio': 0}]

                clientes_con_productos.append({
                    'cliente': cliente,
                    'productos': productos_con_precio
                })

            def enviar_correo(data):
                """Función para enviar correo en thread"""
                cliente = data['cliente']
                productos = data['productos']
                try:
                    result = gmail.send_offer_email(
                        cliente_nombre=cliente.nombre,
                        cliente_email=cliente.correo,
                        productos=productos,
                        descuento_porcentaje=descuento
                    )
                    if result:
                        # Marcar cliente como que recibió oferta
                        Cliente.objects.filter(pk=cliente.pk).update(ultima_oferta_enviada=timezone.now())
                    return {'success': result is not None, 'cliente': cliente.nombre, 'cliente_id': cliente.pk}
                except Exception as e:
                    return {'success': False, 'cliente': cliente.nombre, 'cliente_id': cliente.pk, 'error': str(e)}

            # Enviar correos en paralelo (5 threads para no saturar SMTP)
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(enviar_correo, data): data for data in clientes_con_productos}

                for future in as_completed(futures):
                    result = future.result()
                    with lock:
                        if result['success']:
                            enviados += 1
                        else:
                            errores.append(f"Error enviando a {result['cliente']}")

                    # Enviar progreso
                    yield f"data: {json.dumps({'type': 'progress', 'enviados': enviados, 'total': total, 'current': enviados + len(errores), 'cliente': result['cliente']})}\n\n"

            # Enviar evento final
            yield f"data: {json.dumps({'type': 'complete', 'enviados': enviados, 'total': total, 'errores': errores[:10]})}\n\n"

        response = StreamingHttpResponse(
            generate_progress(),
            content_type='text/event-stream'
        )
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        # Headers CORS para SSE
        response['Access-Control-Allow-Origin'] = '*'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

class TransaccionViewSet(viewsets.ModelViewSet):
    queryset = Transaccion.objects.all()
    serializer_class = TransaccionSerializer

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Estadísticas de transacciones"""
        total = Transaccion.objects.count()
        return Response({'total': total})


class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.select_related('proveedor_principal', 'categoria').all()
    serializer_class = ProductoSerializer
    pagination_class = OfertasPagination  # 50 productos por página
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['codigo', 'nombre', 'descripcion', 'stock_actual', 'stock_minimo', 'precio_venta']
    ordering = ['codigo']

    def get_serializer_class(self):
        """
        Usa serializer LIGERO para listas (sin cálculos ML pesados).
        Usa serializer COMPLETO solo para detalle de un producto.
        """
        if self.action == 'list':
            return ProductoListSerializer
        return ProductoSerializer

    def get_queryset(self):
        """Query optimizado con filtros"""
        queryset = Producto.objects.select_related(
            'proveedor_principal', 'categoria'
        ).only(
            'id', 'codigo', 'nombre', 'descripcion',
            'stock_actual', 'stock_minimo',
            'precio_costo', 'precio_venta',
            'activo', 'fecha_registro',
            'proveedor_principal__nombre',
            'categoria__nombre'
        )

        # FILTRO PRINCIPAL: Solo productos activos
        queryset = queryset.filter(activo=True)

        # Filtro de búsqueda
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(codigo__icontains=search) |
                Q(nombre__icontains=search) |
                Q(descripcion__icontains=search) |
                Q(categoria__nombre__icontains=search)
            )

        # Filtro de stock - comparación simple: stock_actual vs stock_minimo
        # Esto debe coincidir con la lógica del badge en el frontend
        filtro_stock = self.request.query_params.get('filtro_stock', None)
        if filtro_stock == 'bajo':
            # Bajo stock: stock_actual < stock_minimo (crítico o bajo)
            # Incluye crítico (<50%) y bajo (<100%)
            queryset = queryset.filter(stock_actual__lt=F('stock_minimo'))
        elif filtro_stock == 'normal':
            # Stock normal: stock_actual >= stock_minimo
            queryset = queryset.filter(stock_actual__gte=F('stock_minimo'))

        return queryset

    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """Productos con stock bajo"""
        productos = Producto.objects.filter(
            stock_actual__lt=F('stock_minimo')
        )
        serializer = self.get_serializer(productos, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def top_selling(self, request):
        """Top productos más vendidos"""
        limit = int(request.query_params.get('limit', 10))

        # Obtener top productos desde DetalleVenta
        top_productos = DetalleVenta.objects.values('producto__codigo', 'producto__descripcion').annotate(
            total_vendido=Sum('cantidad'),
            total_ventas=Sum('subtotal')
        ).order_by('-total_vendido')[:limit]

        return Response(top_productos)

    @action(detail=False, methods=['get'])
    def criticos(self, request):
        """
        Productos críticos ordenados por prioridad de reposición.

        Retorna productos con nivel de riesgo crítico o alto, ordenados por:
        1. Días de cobertura (menor a mayor) - los que se agotan primero
        2. Sin datos pero con stock crítico

        Query params:
            - limit: Número máximo de productos (default: 100)
        """
        from .stock_service import stock_service

        limit = int(request.query_params.get('limit', 100))

        # Obtener productos activos
        productos = Producto.objects.filter(activo=True).select_related(
            'proveedor_principal', 'categoria'
        )[:min(limit * 3, 1000)]  # Procesar más para filtrar después

        criticos = []

        for producto in productos:
            try:
                # Obtener métricas del producto
                metricas = stock_service.obtener_metricas_producto(producto)
                nivel_riesgo = metricas.get('nivel_riesgo', 'desconocido')

                # Filtrar solo productos críticos o de alto riesgo
                if nivel_riesgo in ['critico', 'alto', 'sin_datos_stock_critico']:
                    stock_calc = stock_service.calcular_stock_minimo(producto)
                    dias_cob = metricas.get('dias_cobertura')

                    # Calcular cantidad sugerida para reposición
                    cantidad_sugerida = max(0, stock_calc - producto.stock_actual)

                    criticos.append({
                        'id': producto.id,
                        'codigo': producto.codigo,
                        'nombre': producto.nombre,
                        'descripcion': producto.descripcion,
                        'stock_actual': producto.stock_actual,
                        'stock_minimo_calculado': stock_calc,
                        'cantidad_sugerida': cantidad_sugerida,
                        'dias_cobertura': round(dias_cob, 1) if dias_cob is not None else None,
                        'nivel_riesgo': nivel_riesgo,
                        'demanda_promedio_diaria': metricas.get('demanda_promedio_diaria', 0),
                        'fuente_datos': metricas.get('fuente_datos', 'desconocido'),
                        'requiere_revision': metricas.get('requiere_revision', False),
                        'proveedor': producto.proveedor_principal.nombre if producto.proveedor_principal else 'Sin proveedor',
                        'categoria': producto.categoria.nombre if producto.categoria else 'Sin categoría',
                        'precio_costo': float(producto.precio_costo) if producto.precio_costo else 0,
                    })

                    # Si ya tenemos suficientes, salir
                    if len(criticos) >= limit:
                        break

            except Exception as e:
                print(f"Error procesando producto {producto.codigo}: {str(e)}")
                continue

        # Ordenar por prioridad de criticidad
        # 1. Primero los que tienen días de cobertura (por días asc)
        # 2. Luego los sin datos pero con stock crítico
        criticos_ordenados = sorted(
            criticos,
            key=lambda x: (
                0 if x['dias_cobertura'] is not None else 1,  # Con días primero
                x['dias_cobertura'] if x['dias_cobertura'] is not None else 999,  # Menor días primero
                x['stock_actual']  # Si no hay días, menor stock primero
            )
        )

        return Response({
            'total': len(criticos_ordenados),
            'productos': criticos_ordenados[:limit],
            'timestamp': timezone.now().isoformat(),
            'niveles_riesgo_incluidos': ['critico', 'alto', 'sin_datos_stock_critico'],
            'descripcion': {
                'critico': '< 7 días de cobertura',
                'alto': '7-13 días de cobertura',
                'sin_datos_stock_critico': 'Sin ventas históricas, stock < 5 unidades'
            }
        })

    @action(detail=False, methods=['get'], url_path='ultima-carga')
    def ultima_carga(self, request):
        """Obtiene la fecha y hora de la última carga de inventario"""
        # Buscar el producto más reciente (por fecha de registro)
        producto_reciente = Producto.objects.filter(
            activo=True
        ).order_by('-fecha_registro').first()

        if producto_reciente:
            return Response({
                'fecha_ultima_carga': producto_reciente.fecha_registro.isoformat(),
                'tiene_inventario': True
            })
        else:
            return Response({
                'fecha_ultima_carga': None,
                'tiene_inventario': False
            })

    @action(detail=False, methods=['post'])
    def cargar_excel(self, request):
        """Cargar productos desde archivo Excel"""
        archivo = request.FILES.get('archivo')

        if not archivo:
            return Response(
                {'error': 'No se proporcionó archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar extensión
        if not archivo.name.endswith(('.xlsx', '.xls', '.XLS', '.XLSX')):
            return Response(
                {'error': 'El archivo debe ser Excel (.xlsx o .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from core.parsers.product_excel_parser import ProductExcelParser

            parser = ProductExcelParser(archivo)
            result = parser.parse_and_load()

            return Response({
                'message': 'Archivo procesado correctamente',
                'productos_insertados': result.get('insertados', 0),
                'productos_actualizados': result.get('actualizados', 0),
                'errores': result.get('errores', [])
            })

        except Exception as e:
            return Response(
                {'error': f'Error procesando archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ProveedorViewSet(viewsets.ModelViewSet):
    queryset = Proveedor.objects.all()
    serializer_class = ProveedorSerializer


class OfertaLaboratorioViewSet(viewsets.ModelViewSet):
    queryset = OfertaLaboratorio.objects.filter(activa=True)
    serializer_class = OfertaLaboratorioSerializer

    @action(detail=False, methods=['get'])
    def por_laboratorio(self, request):
        """
        Obtiene ofertas agrupadas por laboratorio con paginación backend.
        Query params:
            - laboratorio: Filtrar por nombre de laboratorio (opcional)
            - proveedor: Filtrar por nombre de proveedor (opcional)
            - activas: Filtrar solo activas (default: true)
            - page: Número de página (default: 1)
            - page_size: Items por página (default: 50, max: 100)
            - search: Búsqueda en código, descripción, laboratorio
        """
        import logging
        logger = logging.getLogger(__name__)

        laboratorio = request.query_params.get('laboratorio', None)
        proveedor = request.query_params.get('proveedor', None)
        solo_activas = request.query_params.get('activas', 'false').lower() == 'true'
        search = request.query_params.get('search', None)

        logger.info(f"📋 Filtros recibidos: laboratorio={laboratorio}, proveedor={proveedor}, activas={solo_activas}, search={search}")

        # Query optimizado con select_related para reducir queries
        ofertas = OfertaLaboratorio.objects.select_related(
            'producto_catalogo',
            'producto_catalogo__proveedor',
            'laboratorio'
        ).only(
            # Solo campos necesarios para reducir payload
            'id', 'precio_normal', 'precio_oferta',
            'descuento', 'fecha_inicio', 'fecha_fin', 'activa', 'created_at',
            'producto_catalogo__id', 'producto_catalogo__codigo', 'producto_catalogo__nombre',
            'producto_catalogo__descripcion', 'producto_catalogo__proveedor__nombre',
            'laboratorio__id', 'laboratorio__nombre'
        )

        total_inicial = ofertas.count()
        logger.info(f"📊 Total ofertas (sin filtros): {total_inicial}")

        # Filtros
        if solo_activas:
            # Filtrar solo ofertas vigentes
            today = timezone.now().date()
            ofertas = ofertas.filter(activa=True, fecha_fin__gte=today)
            logger.info(f"📊 Después de filtro activas (fecha_fin >= {today}): {ofertas.count()}")

        if laboratorio:
            ofertas = ofertas.filter(laboratorio__nombre__icontains=laboratorio)
            logger.info(f"📊 Después de filtro laboratorio '{laboratorio}': {ofertas.count()}")

        if proveedor:
            # Debug: mostrar proveedores disponibles antes del filtro
            proveedores_disponibles = ofertas.values_list('producto_catalogo__proveedor__nombre', flat=True).distinct()
            logger.info(f"🔍 Proveedores disponibles: {list(proveedores_disponibles)}")
            logger.info(f"🔍 Buscando proveedor: '{proveedor}'")

            ofertas = ofertas.filter(producto_catalogo__proveedor__nombre__icontains=proveedor)
            logger.info(f"📊 Después de filtro proveedor '{proveedor}': {ofertas.count()}")

        if search:
            ofertas = ofertas.filter(
                Q(producto_catalogo__codigo__icontains=search) |
                Q(producto_catalogo__nombre__icontains=search) |
                Q(laboratorio__nombre__icontains=search) |
                Q(producto_catalogo__proveedor__nombre__icontains=search)
            )

        # Ordenamiento dinámico (si se especifica)
        ordering = request.query_params.get('ordering', None)
        if ordering:
            # Soportar ordenamiento descendente con prefijo '-'
            ofertas = ofertas.order_by(ordering)
        else:
            # Ordenar por defecto por laboratorio y producto
            ofertas = ofertas.order_by('laboratorio__nombre', 'producto_catalogo__nombre')

        # Aplicar paginación backend
        paginator = OfertasPagination()
        page = paginator.paginate_queryset(ofertas, request)

        if page is not None:
            serializer = OfertaLaboratorioDetalladaSerializer(page, many=True)
            return paginator.get_paginated_response(serializer.data)

        # Fallback sin paginación (no debería llegar aquí normalmente)
        serializer = OfertaLaboratorioDetalladaSerializer(ofertas, many=True)
        return Response({
            'total': ofertas.count(),
            'ofertas': serializer.data
        })

    @action(detail=False, methods=['get'])
    def laboratorios(self, request):
        """
        Lista todos los laboratorios disponibles con ofertas y conteo.
        Devuelve el NOMBRE del laboratorio (no el ID) para mostrar en el filtro.
        """
        laboratorios = (
            OfertaLaboratorio.objects
            .filter(laboratorio__isnull=False)
            .values('laboratorio__nombre')
            .annotate(
                total_ofertas=Count('id'),
                promedio_descuento=Sum('descuento') / Count('id')
            )
            .order_by('-total_ofertas')
        )

        # Mapear laboratorio__nombre a "laboratorio" para mantener compatibilidad con frontend
        laboratorios_list = [
            {
                'laboratorio': lab['laboratorio__nombre'],
                'total_ofertas': lab['total_ofertas'],
                'promedio_descuento': lab['promedio_descuento']
            }
            for lab in laboratorios
        ]

        return Response({
            'total_laboratorios': len(laboratorios_list),
            'laboratorios': laboratorios_list
        })

    @action(detail=False, methods=['get'])
    def proveedores(self, request):
        """
        Lista todos los proveedores disponibles con ofertas y conteo.
        Devuelve el NOMBRE del proveedor para mostrar en el filtro.
        """
        proveedores = (
            OfertaLaboratorio.objects
            .filter(producto_catalogo__proveedor__isnull=False)
            .values('producto_catalogo__proveedor__nombre')
            .annotate(
                total_ofertas=Count('id'),
                promedio_descuento=Sum('descuento') / Count('id')
            )
            .order_by('-total_ofertas')
        )

        # Mapear a formato consistente
        proveedores_list = [
            {
                'proveedor': prov['producto_catalogo__proveedor__nombre'],
                'total_ofertas': prov['total_ofertas'],
                'promedio_descuento': prov['promedio_descuento']
            }
            for prov in proveedores
        ]

        return Response({
            'total_proveedores': len(proveedores_list),
            'proveedores': proveedores_list
        })

    @action(detail=False, methods=['post'])
    def procesar(self, request):
        """Procesar archivo de ofertas (Excel/PDF)"""
        archivo = request.FILES.get('archivo')

        if not archivo:
            return Response(
                {'error': 'No se proporcionó archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # TODO: Implementar lógica de ETL
        # Por ahora solo retorna mensaje de éxito
        return Response({
            'message': f'Archivo {archivo.name} recibido y procesado correctamente',
            'ofertas_creadas': 0
        })


class VentaViewSet(viewsets.ModelViewSet):
    queryset = Venta.objects.select_related('cliente').prefetch_related('detalles__producto').all()
    serializer_class = VentaSerializer
    pagination_class = OfertasPagination  # 50 ventas por página
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['numero', 'cliente__rut', 'cliente__nombre', 'detalles__producto__codigo', 'detalles__producto__descripcion']
    ordering_fields = ['id', 'numero', 'fecha', 'total', 'cliente__rut', 'cliente__nombre']
    ordering = ['-fecha']  # Ordenamiento por defecto

    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener término de búsqueda y normalizarlo (quitar puntos y guiones para RUT)
        search = self.request.query_params.get('search', '')
        if search:
            # Normalizar RUT: quitar puntos y guiones
            search_normalized = search.replace('.', '').replace('-', '').strip()

            # Buscar con ambos formatos (original y normalizado)
            queryset = queryset.filter(
                Q(numero__icontains=search) |
                Q(cliente__rut__icontains=search_normalized) |
                Q(cliente__nombre__icontains=search) |
                Q(detalles__producto__codigo__icontains=search) |
                Q(detalles__producto__descripcion__icontains=search)
            ).distinct()

        return queryset

    def list(self, request, *args, **kwargs):
        # Si hay búsqueda personalizada, no usar el SearchFilter predeterminado
        if request.query_params.get('search'):
            # Remover temporalmente SearchFilter para usar nuestra lógica
            self.filter_backends = [filters.OrderingFilter]
        return super().list(request, *args, **kwargs)

    def get_queryset(self):
        """Query con filtros de búsqueda"""
        queryset = Venta.objects.select_related(
            'cliente', 'metodo_pago', 'estado'
        ).prefetch_related('detalles__producto')

        # Filtro de búsqueda
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(numero__icontains=search) |
                Q(cliente__nombre__icontains=search) |
                Q(cliente__rut__icontains=search) |
                Q(detalles__producto__codigo__icontains=search) |
                Q(detalles__producto__nombre__icontains=search)
            ).distinct()

        # Filtro por fecha
        fecha_desde = self.request.query_params.get('fecha_desde', None)
        fecha_hasta = self.request.query_params.get('fecha_hasta', None)
        if fecha_desde:
            queryset = queryset.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha__lte=fecha_hasta)

        # Filtro por cliente
        cliente_id = self.request.query_params.get('cliente', None)
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)

        return queryset.order_by('-fecha', '-id')

    @action(detail=False, methods=['get'], url_path='ultima-carga')
    def ultima_carga(self, request):
        """
        Obtiene la fecha y hora de la última carga/importación de ventas.
        Usa tabla de configuración si existe, sino la fecha de venta más reciente.
        """
        from django.db import connection

        # Intentar obtener fecha de última importación desde configuración
        fecha_ultima_importacion = None
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT valor FROM core_configuracion WHERE clave = 'ultima_carga_ventas'"
                )
                row = cursor.fetchone()
                if row:
                    fecha_ultima_importacion = row[0]
        except Exception:
            # Tabla no existe o error - usar fallback
            pass

        # Fallback: usar fecha de venta más reciente si no hay config
        if not fecha_ultima_importacion:
            venta_reciente = Venta.objects.filter(
                estado__codigo='completada'
            ).order_by('-fecha').first()
            if venta_reciente:
                fecha_ultima_importacion = venta_reciente.fecha.isoformat()

        tiene_ventas = Venta.objects.filter(estado__codigo='completada').exists()

        return Response({
            'fecha_ultima_carga': fecha_ultima_importacion,
            'tiene_ventas': tiene_ventas
        })

    @action(detail=False, methods=['post'])
    def cargar_excel(self, request):
        """Cargar ventas desde archivo Excel"""
        archivo = request.FILES.get('archivo')

        if not archivo:
            return Response(
                {'error': 'No se proporcionó archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar extensión
        if not archivo.name.endswith(('.xlsx', '.xls', '.XLS', '.XLSX')):
            return Response(
                {'error': 'El archivo debe ser Excel (.xlsx o .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from core.parsers.sales_excel_parser import SalesExcelParser
            from django.db import connection

            parser = SalesExcelParser(archivo)
            result = parser.parse_and_load()

            # Registrar fecha/hora de carga exitosa en configuración (si la tabla existe)
            if result.get('ventas_insertadas', 0) > 0:
                try:
                    fecha_carga = timezone.now().isoformat()
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            INSERT INTO core_configuracion (clave, valor, actualizado_en)
                            VALUES ('ultima_carga_ventas', %s, NOW())
                            ON CONFLICT (clave) DO UPDATE SET valor = %s, actualizado_en = NOW()
                        """, [fecha_carga, fecha_carga])
                except Exception:
                    # Tabla no existe - ignorar silenciosamente
                    pass

            return Response({
                'message': 'Archivo de ventas procesado correctamente',
                'ventas_insertadas': result.get('ventas_insertadas', 0),
                'ventas_duplicadas': result.get('ventas_duplicadas', 0),
                'detalles_insertados': result.get('detalles_insertados', 0),
                'detalles_duplicados_omitidos': result.get('detalles_duplicados_omitidos', 0),
                'clientes_creados': result.get('clientes_creados', 0),
                'productos_no_encontrados': result.get('productos_no_encontrados', 0),
                'errores': result.get('errores', [])[:20]  # Mostrar solo primeros 20 errores
            })

        except Exception as e:
            return Response(
                {'error': f'Error procesando archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DashboardViewSet(viewsets.GenericViewSet):
    """Endpoints específicos para el dashboard"""
    # Queryset vacío para que el router funcione correctamente
    queryset = Venta.objects.none()

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Estadísticas generales del dashboard"""
        # Total ventas
        total_ventas = Venta.objects.filter(estado__codigo='completada').aggregate(
            total=Sum('total')
        )['total'] or 0

        # Ventas del mes más reciente (último mes con ventas)
        # Obtener la venta más reciente
        venta_reciente = Venta.objects.filter(estado__codigo='completada').order_by('-fecha').first()

        if venta_reciente:
            # Calcular el primer día del mes de esa venta
            mes_reciente = venta_reciente.fecha.replace(day=1)
            # Calcular el último día del mes
            if mes_reciente.month == 12:
                fin_mes = mes_reciente.replace(year=mes_reciente.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                fin_mes = mes_reciente.replace(month=mes_reciente.month + 1, day=1) - timedelta(days=1)

            ventas_mes = Venta.objects.filter(
                estado__codigo='completada',
                fecha__gte=mes_reciente,
                fecha__lte=fin_mes
            ).aggregate(total=Sum('total'))['total'] or 0
        else:
            ventas_mes = 0

        # Productos en stock
        productos_stock = Producto.objects.filter(activo=True).count()

        # Clientes activos (con compras en los últimos 6 meses)
        hace_6_meses = timezone.now() - timedelta(days=180)
        clientes_activos = Cliente.objects.filter(
            ventas__fecha__gte=hace_6_meses
        ).distinct().count()

        return Response({
            'total_ventas': total_ventas,
            'ventas_mes': ventas_mes,
            'productos_stock': productos_stock,
            'clientes_activos': clientes_activos
        })

    @action(detail=False, methods=['get'])
    def sales(self, request):
        """Datos de ventas para gráficos"""
        # Ventas de los últimos 12 meses
        hace_12_meses = timezone.now() - timedelta(days=365)

        # Usar TruncMonth para PostgreSQL (compatible con todas las BD)
        ventas_mensuales = Venta.objects.filter(
            estado__codigo='completada',
            fecha__gte=hace_12_meses
        ).annotate(
            mes=TruncMonth('fecha')
        ).values('mes').annotate(
            total=Sum('total')
        ).order_by('mes')

        # Formatear la respuesta
        result = []
        for venta in ventas_mensuales:
            result.append({
                'mes': venta['mes'].strftime('%Y-%m') if venta['mes'] else None,
                'total': float(venta['total']) if venta['total'] else 0
            })

        return Response(result)

    @action(detail=False, methods=['get'], url_path='top-products')
    def top_products(self, request):
        """Top 10 productos más vendidos"""
        limit = int(request.query_params.get('limit', 10))

        top_productos = DetalleVenta.objects.values(
            'producto__codigo',
            'producto__nombre'
        ).annotate(
            cantidad=Sum('cantidad'),
            ventas=Sum('subtotal')
        ).order_by('-cantidad')[:limit]

        return Response(list(top_productos))


class SugerenciaCompraViewSet(viewsets.ModelViewSet):
    """
    ViewSet para Sugerencias de Compra Inteligentes.

    Endpoints:
    - /api/sugerencias-compra/ - CRUD básico
    - /api/sugerencias-compra/generar/ - Genera sugerencias desde productos críticos
    - /api/sugerencias-compra/consolidar/ - Consolida por proveedor
    - /api/sugerencias-compra/export-excel/ - Exporta a Excel
    """
    queryset = SugerenciaCompra.objects.all()
    serializer_class = SugerenciaCompraSerializer
    pagination_class = OfertasPagination

    def get_queryset(self):
        """Filtra sugerencias"""
        queryset = super().get_queryset()

        # Filtro por estado
        procesada = self.request.query_params.get('procesada', None)
        if procesada is not None:
            queryset = queryset.filter(procesada=procesada.lower() == 'true')

        # Filtro por proveedor
        proveedor = self.request.query_params.get('proveedor', None)
        if proveedor:
            queryset = queryset.filter(proveedor_recomendado__id=proveedor)

        return queryset.select_related('producto', 'proveedor_recomendado').order_by('-prioridad', '-fecha_creacion')

    @action(detail=False, methods=['get'])
    def todas(self, request):
        """
        Endpoint unificado optimizado con paginación.

        Query params:
            - tipo: 'bajo_stock', 'estacional', 'epidemiologico', 'all' (default: 'all')
            - page: número de página (default: 1)
            - page_size: items por página (default: 50, max: 100)

        Returns:
            {
                'results': [...],
                'pagination': {
                    'page': 1,
                    'page_size': 50,
                    'total_items': 150,
                    'total_pages': 3,
                    'has_next': true,
                    'has_previous': false
                },
                'conteos': {
                    'bajo_stock': 50,
                    'estacional': 30,
                    'epidemiologico': 10,
                    'total': 90
                },
                'timestamp': '...'
            }
        """
        tipo = request.query_params.get('tipo', 'all')
        page = int(request.query_params.get('page', 1))
        page_size = min(int(request.query_params.get('page_size', 50)), 100)

        # Query base optimizado
        base_query = SugerenciaCompra.objects.filter(
            procesada=False
        ).select_related(
            'producto',
            'producto__categoria',
            'proveedor_recomendado'
        ).only(
            'id', 'tipo', 'cantidad_sugerida', 'prioridad', 'razon',
            'precio_unitario', 'tiene_oferta', 'precio_oferta', 'descuento_porcentaje',
            'confianza_ml', 'fuente_datos', 'dias_cobertura', 'fecha_creacion',
            'producto__id', 'producto__codigo', 'producto__nombre', 'producto__descripcion',
            'producto__stock_actual', 'producto__stock_minimo', 'producto__precio_costo',
            'producto__categoria__nombre',
            'proveedor_recomendado__id', 'proveedor_recomendado__nombre'
        )

        # Conteos rápidos (una sola query con agregación)
        conteos = SugerenciaCompra.objects.filter(procesada=False).aggregate(
            bajo_stock=Count('id', filter=Q(tipo='bajo_stock')),
            estacional=Count('id', filter=Q(tipo='estacional')),
            epidemiologico=Count('id', filter=Q(tipo='epidemiologico')),
            total=Count('id')
        )

        # Filtrar por tipo si se especifica
        if tipo != 'all':
            queryset = base_query.filter(tipo=tipo)
        else:
            queryset = base_query

        # Ordenar según tipo
        if tipo == 'estacional':
            queryset = queryset.order_by('-confianza_ml', '-prioridad')
        else:
            queryset = queryset.order_by('-prioridad', '-fecha_creacion')

        # Paginación manual eficiente
        total_items = queryset.count()
        total_pages = (total_items + page_size - 1) // page_size
        start = (page - 1) * page_size
        end = start + page_size

        # Obtener solo la página actual
        sugerencias = list(queryset[start:end])
        serializer = self.get_serializer(sugerencias, many=True)

        return Response({
            'results': serializer.data,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_items': total_items,
                'total_pages': total_pages,
                'has_next': page < total_pages,
                'has_previous': page > 1
            },
            'conteos': conteos,
            'timestamp': timezone.now().isoformat()
        })

    @action(detail=False, methods=['get'], url_path='low-stock')
    def low_stock(self, request):
        """
        Sugerencias por bajo stock (optimizado)
        Límite: 50 sugerencias por defecto
        """
        limite = int(request.query_params.get('limite', 50))

        sugerencias = SugerenciaCompra.objects.filter(
            tipo='bajo_stock',
            procesada=False
        ).select_related(
            'producto',
            'producto__categoria',
            'proveedor_recomendado'
        ).order_by('-prioridad', '-fecha_creacion')[:limite]

        serializer = self.get_serializer(sugerencias, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def season(self, request):
        """
        Sugerencias estacionales (optimizado)
        Límite: 50 sugerencias por defecto
        """
        limite = int(request.query_params.get('limite', 50))

        sugerencias = SugerenciaCompra.objects.filter(
            tipo='estacional',
            procesada=False
        ).select_related(
            'producto',
            'producto__categoria',
            'proveedor_recomendado'
        ).order_by('-confianza_ml', '-prioridad')[:limite]

        serializer = self.get_serializer(sugerencias, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def epidemiological(self, request):
        """
        Sugerencias epidemiológicas (optimizado)
        Límite: 50 sugerencias por defecto
        """
        limite = int(request.query_params.get('limite', 50))

        sugerencias = SugerenciaCompra.objects.filter(
            tipo='epidemiologico',
            procesada=False
        ).select_related(
            'producto',
            'producto__categoria',
            'proveedor_recomendado'
        ).order_by('-prioridad')[:limite]

        serializer = self.get_serializer(sugerencias, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def generar(self, request):
        """
        Genera sugerencias de compra desde productos críticos.

        Query params:
            - limite: Número máximo de productos a procesar (default: 100)
            - forzar_mapeo: Si True, mapea automáticamente productos sin mapping (default: true)
        """
        from .purchase_optimizer_service import purchase_optimizer

        limite = int(request.query_params.get('limite', 100))
        forzar_mapeo = request.query_params.get('forzar_mapeo', 'true').lower() == 'true'

        try:
            stats = purchase_optimizer.generar_sugerencias_desde_criticos(
                limite=limite,
                forzar_mapeo=forzar_mapeo
            )

            return Response({
                'success': True,
                'mensaje': f'Sugerencias generadas correctamente',
                'estadisticas': stats,
                'timestamp': timezone.now().isoformat()
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def consolidar(self, request):
        """
        Consolida sugerencias por proveedor (optimizado).

        Agrupa sugerencias no procesadas por proveedor,
        calcula totales y valida mínimos de pedido.

        Returns:
            {
                'MEDIVEN': {
                    'total': 150000,
                    'cumple_minimo': true,
                    'minimo_requerido': 50000,
                    'sugerencias': [...]
                },
                'SOCOFAR': {...}
            }
        """
        from .purchase_optimizer_service import purchase_optimizer

        try:
            # Verificar rápidamente si hay sugerencias
            if not SugerenciaCompra.objects.filter(procesada=False, incluida_en_orden=False).exists():
                return Response({
                    'success': True,
                    'consolidado': {},
                    'total_proveedores': 0,
                    'timestamp': timezone.now().isoformat(),
                    'mensaje': 'No hay sugerencias pendientes para consolidar'
                })

            consolidado = purchase_optimizer.consolidar_por_proveedor(incluir_solo_activas=True)

            # Limitar sugerencias en respuesta a 50 por proveedor para reducir payload
            limite_sugerencias = 50

            # Serializar para respuesta
            resultado = {}
            for prov_nombre, data in consolidado.items():
                # Limitar sugerencias serializadas
                sugerencias_limitadas = data['sugerencias'][:limite_sugerencias]
                sugerencias_serialized = SugerenciaCompraSerializer(
                    sugerencias_limitadas, many=True
                ).data

                resultado[prov_nombre] = {
                    'proveedor': {
                        'id': data['proveedor'].id,
                        'nombre': data['proveedor'].nombre
                    },
                    'total': float(data['total']),
                    'cumple_minimo': data['cumple_minimo'],
                    'minimo_requerido': float(data['minimo_requerido']),
                    'cantidad_productos': len(data['sugerencias']),
                    'sugerencias': sugerencias_serialized,
                    'total_sugerencias': len(data['sugerencias']),
                    'sugerencias_mostradas': len(sugerencias_limitadas)
                }

            return Response({
                'success': True,
                'consolidado': resultado,
                'total_proveedores': len(resultado),
                'timestamp': timezone.now().isoformat()
            })

        except Exception as e:
            print(f"Error en consolidar: {str(e)}")
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Exporta sugerencias consolidadas a Excel.

        Query params:
            - proveedor: Filtrar por proveedor específico (opcional)
        """
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from .purchase_optimizer_service import purchase_optimizer

        try:
            consolidado = purchase_optimizer.consolidar_por_proveedor(incluir_solo_activas=True)

            # Filtrar por proveedor si se especifica
            proveedor_filtro = request.query_params.get('proveedor', None)
            if proveedor_filtro:
                consolidado = {
                    k: v for k, v in consolidado.items()
                    if k.upper() == proveedor_filtro.upper()
                }

            # Crear libro Excel
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Eliminar hoja por defecto

            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            title_font = Font(size=14, bold=True)

            # Crear hoja por proveedor
            for prov_nombre, data in consolidado.items():
                ws = wb.create_sheet(title=prov_nombre[:31])  # Excel limita a 31 chars

                # Título
                ws['A1'] = f"ORDEN DE COMPRA - {prov_nombre}"
                ws['A1'].font = title_font
                ws.merge_cells('A1:H1')

                # Info del proveedor
                ws['A2'] = f"Fecha: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
                ws['A3'] = f"Total: ${data['total']:,.0f}"
                ws['A4'] = f"Mínimo Requerido: ${data['minimo_requerido']:,.0f}"
                ws['A5'] = f"Cumple Mínimo: {'SÍ' if data['cumple_minimo'] else 'NO'}"

                # Headers
                row = 7
                headers = ['Código', 'Producto', 'Cantidad', 'Precio Unit.', 'Subtotal', 'Tiene Oferta', 'Descuento %', 'Ahorro']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                # Datos
                row = 8
                for sug in data['sugerencias']:
                    ws.cell(row=row, column=1, value=sug.producto.codigo)
                    ws.cell(row=row, column=2, value=sug.producto.nombre)
                    ws.cell(row=row, column=3, value=sug.cantidad_sugerida)
                    ws.cell(row=row, column=4, value=float(sug.precio_unitario) if sug.precio_unitario else 0)
                    ws.cell(row=row, column=5, value=sug.total)
                    ws.cell(row=row, column=6, value='SÍ' if sug.tiene_oferta else 'NO')
                    ws.cell(row=row, column=7, value=float(sug.descuento_porcentaje) if sug.descuento_porcentaje else 0)
                    ws.cell(row=row, column=8, value=sug.ahorro_total)
                    row += 1

                # Totales
                row += 1
                ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
                ws.cell(row=row, column=5, value=float(data['total'])).font = Font(bold=True)

                # Ajustar anchos
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 50
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 12

            # Generar respuesta HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"orden_compra_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename={filename}'

            wb.save(response)
            return response

        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def rapido(self, request):
        """
        Endpoint RÁPIDO para sugerencias de compra.

        Criterios:
        1. Productos con stock < stock_minimo Y demanda reciente (ventas últimos 90 días)
        2. Busca ofertas por mapping existente O por similitud de nombre (fuzzy matching)
        3. EXCLUYE productos sin ventas en 1 año (probable obsolescencia)
        4. Incluye mejor oferta disponible para cada producto

        Query params:
            - limite: máximo de sugerencias (default: 50)
        """
        from django.db.models import OuterRef, Subquery, Exists
        from decimal import Decimal
        from fuzzywuzzy import fuzz
        from .ml_service import seasonal_service

        limite = int(request.query_params.get('limite', 50))
        hoy = timezone.now().date()
        hace_90_dias = timezone.now() - timedelta(days=90)

        # Mes próximo para predicción estacional
        mes_proximo = hoy.month + 1 if hoy.month < 12 else 1
        año_proximo = hoy.year if hoy.month < 12 else hoy.year + 1

        try:
            # ============================================================
            # PASO 1: Productos con demanda reciente (ventas últimos 90 días)
            # ============================================================
            productos_con_demanda = DetalleVenta.objects.filter(
                venta__fecha__gte=hace_90_dias,
                venta__estado__codigo='completada'
            ).values('producto_id').annotate(
                total_vendido=Sum('cantidad'),
                ultima_venta=Max('venta__fecha')
            )

            productos_ids_con_demanda = set(
                p['producto_id'] for p in productos_con_demanda
            )

            # Mapa de ventas por producto
            ventas_por_producto = {
                p['producto_id']: {
                    'total_vendido': p['total_vendido'],
                    'ultima_venta': p['ultima_venta']
                }
                for p in productos_con_demanda
            }

            # ============================================================
            # PASO 2: Productos con bajo stock Y demanda
            # ============================================================
            productos_bajo_stock = list(Producto.objects.filter(
                activo=True,
                id__in=productos_ids_con_demanda,
                stock_actual__lt=F('stock_minimo')
            ).select_related('categoria', 'proveedor_principal')[:limite])

            # ============================================================
            # PASO 3: Obtener TODAS las ofertas activas
            # ============================================================
            from .models import ProductoCatalogo, ProductoProveedorMapping

            # Cargar todas las ofertas activas en memoria para fuzzy matching
            ofertas_activas = list(OfertaLaboratorio.objects.filter(
                activa=True,
                fecha_inicio__lte=hoy,
                fecha_fin__gte=hoy
            ).select_related('producto_catalogo', 'producto_catalogo__proveedor', 'laboratorio').values(
                'id',
                'producto_catalogo__codigo',
                'producto_catalogo__nombre',
                'producto_catalogo__proveedor__nombre',
                'laboratorio__nombre',
                'precio_normal',
                'precio_oferta',
                'descuento',
                'fecha_fin'
            ))

            # Obtener mappings existentes
            producto_ids = [p.id for p in productos_bajo_stock]
            mappings = ProductoProveedorMapping.objects.filter(
                producto_interno_id__in=producto_ids,
                activo=True
            ).select_related('proveedor')

            # Mapa de mappings por producto
            mappings_por_producto = {}
            for m in mappings:
                if m.producto_interno_id not in mappings_por_producto:
                    mappings_por_producto[m.producto_interno_id] = []
                mappings_por_producto[m.producto_interno_id].append({
                    'codigo': m.codigo_proveedor,
                    'nombre_catalogo': m.nombre_en_catalogo,
                    'proveedor': m.proveedor.nombre if m.proveedor else None,
                    'proveedor_id': m.proveedor_id,
                    'confianza': float(m.confianza)
                })

            # Indexar ofertas por código para búsqueda rápida
            ofertas_por_codigo = {o['producto_catalogo__codigo']: o for o in ofertas_activas if o['producto_catalogo__codigo']}

            # ============================================================
            # FUNCIÓN: Normalizar texto para comparación
            # ============================================================
            def normalizar_nombre(texto):
                if not texto:
                    return ""
                import re
                texto = texto.upper()
                # Normalizar abreviaturas comunes
                reemplazos = {
                    'PARACETA.': 'PARACETAMOL', 'COM.': 'COMP', 'CAP.': 'CAP',
                    'TAB.': 'TAB', 'JBE.': 'JARABE', 'MG,': 'MG ', 'MG.': 'MG ',
                    '(BE)': 'BE', ' X ': ' ', 'X ': ' '
                }
                for old, new in reemplazos.items():
                    texto = texto.replace(old, new)
                texto = re.sub(r'[^A-Z0-9\s]', ' ', texto)
                texto = re.sub(r'\s+', ' ', texto).strip()
                return texto

            # ============================================================
            # FUNCIÓN: Buscar mejor oferta por similitud
            # ============================================================
            def buscar_mejor_oferta(producto, mappings_producto):
                mejor_oferta = None
                mejor_score = 0
                proveedor_nombre = None

                # 1. Primero buscar por mapping existente (más confiable)
                if mappings_producto:
                    for mapping in mappings_producto:
                        if mapping['codigo'] in ofertas_por_codigo:
                            oferta = ofertas_por_codigo[mapping['codigo']]
                            # Score alto porque hay mapping
                            score = 100 + mapping['confianza']
                            if score > mejor_score:
                                mejor_score = score
                                mejor_oferta = oferta
                                proveedor_nombre = mapping['proveedor']

                # 2. Si no hay mapping o no encontró oferta, buscar por similitud de nombre
                if not mejor_oferta and ofertas_activas:
                    nombre_producto = normalizar_nombre(producto.nombre)

                    for oferta in ofertas_activas:
                        nombre_oferta = normalizar_nombre(oferta.get('producto_catalogo__nombre', ''))

                        # Calcular similitud usando fuzzy matching
                        score = fuzz.token_sort_ratio(nombre_producto, nombre_oferta)

                        # Solo considerar si score >= 60 (60% similitud)
                        if score >= 60 and score > mejor_score:
                            mejor_score = score
                            mejor_oferta = oferta
                            proveedor_nombre = oferta.get('producto_catalogo__proveedor__nombre')

                return mejor_oferta, proveedor_nombre, mejor_score

            # ============================================================
            # PASO 4: Construir respuesta
            # ============================================================
            sugerencias = []

            # Cache de predicciones por categoría para evitar cálculos repetidos
            predicciones_por_categoria = {}

            def obtener_prediccion_estacional(categoria_nombre):
                """Obtiene predicción estacional para una categoría (con cache)"""
                if not categoria_nombre:
                    return None
                if categoria_nombre not in predicciones_por_categoria:
                    try:
                        # Obtener predicción base para el mes próximo
                        pred_base = seasonal_service._estimate_default_transactions(mes_proximo, categoria_nombre)
                        # Obtener promedio anual para comparar
                        promedio_anual = sum(
                            seasonal_service._estimate_default_transactions(m, categoria_nombre)
                            for m in range(1, 13)
                        ) / 12
                        # Calcular tendencia
                        if pred_base > promedio_anual * 1.15:
                            tendencia = 'ALTA'
                            factor_ajuste = 1.25  # Aumentar 25% para alta demanda
                        elif pred_base < promedio_anual * 0.85:
                            tendencia = 'BAJA'
                            factor_ajuste = 0.90  # Reducir 10% para baja demanda
                        else:
                            tendencia = 'NORMAL'
                            factor_ajuste = 1.0

                        predicciones_por_categoria[categoria_nombre] = {
                            'prediccion': int(pred_base),
                            'promedio_anual': int(promedio_anual),
                            'tendencia': tendencia,
                            'factor_ajuste': factor_ajuste,
                            'mes_nombre': ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                                          'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes_proximo - 1]
                        }
                    except Exception as e:
                        print(f"⚠️ Error predicción estacional para {categoria_nombre}: {e}")
                        predicciones_por_categoria[categoria_nombre] = None
                return predicciones_por_categoria.get(categoria_nombre)

            for producto in productos_bajo_stock:
                # Calcular cantidad sugerida base
                cantidad_base = max(1, producto.stock_minimo - producto.stock_actual)

                # Obtener predicción estacional para la categoría
                categoria_nombre = producto.categoria.nombre if producto.categoria else None
                pred_estacional = obtener_prediccion_estacional(categoria_nombre)

                # Ajustar cantidad según demanda estacional
                if pred_estacional and pred_estacional['tendencia'] == 'ALTA':
                    cantidad_sugerida = int(cantidad_base * pred_estacional['factor_ajuste'])
                else:
                    cantidad_sugerida = cantidad_base

                # Datos de demanda
                demanda = ventas_por_producto.get(producto.id, {})

                # Buscar mejor oferta (por mapping o similitud)
                mappings_producto = mappings_por_producto.get(producto.id, [])
                mejor_oferta, proveedor_nombre, match_score = buscar_mejor_oferta(
                    producto, mappings_producto
                )

                # Calcular días de cobertura
                total_vendido = demanda.get('total_vendido', 0)
                demanda_diaria = total_vendido / 90 if total_vendido > 0 else 0
                dias_cobertura = round(producto.stock_actual / demanda_diaria, 1) if demanda_diaria > 0 else None

                # Determinar prioridad
                if dias_cobertura is not None:
                    if dias_cobertura < 7:
                        prioridad = 'critica'
                    elif dias_cobertura < 14:
                        prioridad = 'alta'
                    elif dias_cobertura < 30:
                        prioridad = 'media'
                    else:
                        prioridad = 'baja'
                else:
                    prioridad = 'media'

                sugerencia = {
                    'producto': {
                        'id': producto.id,
                        'codigo': producto.codigo,
                        'nombre': producto.nombre,
                        'descripcion': producto.descripcion or producto.nombre,
                        'categoria': producto.categoria.nombre if producto.categoria else None,
                    },
                    'stock': {
                        'actual': producto.stock_actual,
                        'minimo': producto.stock_minimo,
                        'faltante': cantidad_sugerida,
                    },
                    'demanda': {
                        'total_90_dias': total_vendido,
                        'promedio_diario': round(demanda_diaria, 2),
                        'dias_cobertura': dias_cobertura,
                        'ultima_venta': demanda.get('ultima_venta').isoformat() if demanda.get('ultima_venta') else None,
                    },
                    'cantidad_sugerida': cantidad_sugerida,
                    'cantidad_base': cantidad_base,  # Cantidad sin ajuste estacional
                    'prioridad': prioridad,
                    'precio_costo': float(producto.precio_costo) if producto.precio_costo else 0,
                    'oferta': None,
                    'proveedor': proveedor_nombre or (producto.proveedor_principal.nombre if producto.proveedor_principal else None),
                    # Información de demanda estacional
                    'estacional': {
                        'mes_proximo': pred_estacional['mes_nombre'] if pred_estacional else None,
                        'tendencia': pred_estacional['tendencia'] if pred_estacional else None,
                        'prediccion_transacciones': pred_estacional['prediccion'] if pred_estacional else None,
                        'promedio_anual': pred_estacional['promedio_anual'] if pred_estacional else None,
                        'ajuste_aplicado': pred_estacional['tendencia'] == 'ALTA' if pred_estacional else False,
                        'factor_ajuste': pred_estacional['factor_ajuste'] if pred_estacional else 1.0,
                        'cantidad_extra': cantidad_sugerida - cantidad_base if pred_estacional and pred_estacional['tendencia'] == 'ALTA' else 0,
                    } if pred_estacional else None,
                }

                # Agregar oferta si existe
                if mejor_oferta:
                    dias_vigencia = (mejor_oferta['fecha_fin'] - hoy).days
                    precio_lista = float(mejor_oferta['precio_normal']) if mejor_oferta.get('precio_normal') else 0
                    precio_oferta = float(mejor_oferta['precio_oferta']) if mejor_oferta.get('precio_oferta') else 0
                    ahorro_unitario = precio_lista - precio_oferta

                    sugerencia['oferta'] = {
                        'codigo': mejor_oferta.get('producto_catalogo__codigo', ''),
                        'nombre': mejor_oferta.get('producto_catalogo__nombre', ''),
                        'laboratorio': mejor_oferta.get('laboratorio__nombre', ''),
                        'precio_lista': precio_lista,
                        'precio_oferta': precio_oferta,
                        'descuento': float(mejor_oferta['descuento']) if mejor_oferta.get('descuento') else 0,
                        'ahorro_unitario': ahorro_unitario,
                        'ahorro_total': ahorro_unitario * cantidad_sugerida,
                        'dias_vigencia': dias_vigencia,
                        'match_score': match_score,  # Score de coincidencia
                    }
                    sugerencia['proveedor'] = proveedor_nombre or mejor_oferta.get('producto_catalogo__proveedor__nombre', '')

                sugerencias.append(sugerencia)

            # Ordenar por prioridad y luego por demanda
            orden_prioridad = {'critica': 0, 'alta': 1, 'media': 2, 'baja': 3}
            sugerencias.sort(key=lambda x: (
                orden_prioridad.get(x['prioridad'], 99),
                -(x['demanda']['total_90_dias'] or 0)
            ))

            # ============================================================
            # PASO 5: Estadísticas resumen
            # ============================================================
            total_productos = len(sugerencias)
            con_oferta = sum(1 for s in sugerencias if s['oferta'])
            ahorro_total = sum(
                s['oferta']['ahorro_total'] for s in sugerencias if s['oferta']
            )

            return Response({
                'success': True,
                'sugerencias': sugerencias,
                'resumen': {
                    'total_productos': total_productos,
                    'con_oferta': con_oferta,
                    'sin_oferta': total_productos - con_oferta,
                    'ahorro_potencial': round(ahorro_total, 0),
                    'criticos': sum(1 for s in sugerencias if s['prioridad'] == 'critica'),
                    'altos': sum(1 for s in sugerencias if s['prioridad'] == 'alta'),
                },
                'filtros_aplicados': {
                    'demanda_minima': 'ventas últimos 90 días',
                    'excluidos': 'productos sin movimiento',
                    'matching': 'mapping existente + fuzzy matching (60%+ similitud)',
                },
                'timestamp': timezone.now().isoformat()
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
